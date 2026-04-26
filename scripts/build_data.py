"""
Nouveaux Recits d'Entreprise - Import SBTi + ADEME
"""
import json, os, re, sys, hashlib
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

SBTI_LOCAL = "data/companies-excel.xlsx"
ADEME_LOCAL = "data/bilans-ges.xlsx"
OUTPUT_FILE = "data/projects.json"
EDITORIAL_FILE = "data/editorial.json"

SECTOR_MAP = {
    "food and beverage": "Agroalimentaire", "food and staples": "Agroalimentaire",
    "agricultural": "Agroalimentaire", "retailing": "Distribution",
    "textile": "Mode & Textile", "apparel": "Mode & Textile",
    "footwear": "Mode & Textile", "luxury": "Mode & Textile",
    "construction": "BTP & Immobilier", "real estate": "BTP & Immobilier",
    "building products": "BTP & Immobilier", "homebuilding": "BTP & Immobilier",
    "electrical equipment": "Industrie manufacturière", "machinery": "Industrie manufacturière",
    "consumer durables": "Industrie manufacturière", "household": "Industrie manufacturière",
    "containers": "Industrie manufacturière", "packaging": "Industrie manufacturière",
    "forest and paper": "Industrie manufacturière", "paper": "Industrie manufacturière",
    "automobiles": "Transport", "transportation": "Transport",
    "air freight": "Transport", "marine": "Transport", "airlines": "Transport",
    "software": "Tech & Numérique", "technology hardware": "Tech & Numérique",
    "telecommunication": "Tech & Numérique", "media": "Tech & Numérique",
    "semiconductor": "Tech & Numérique", "it services": "Tech & Numérique",
    "banks": "Finance", "financials": "Finance", "insurance": "Finance",
    "capital markets": "Finance", "consumer finance": "Finance",
    "chemicals": "Industrie chimique", "metals": "Industrie lourde",
    "mining": "Industrie lourde", "steel": "Industrie lourde",
    "pharma": "Santé & Pharma", "biotech": "Santé & Pharma",
    "healthcare": "Santé & Pharma", "life sciences": "Santé & Pharma",
    "hotels": "Tourisme & Loisirs", "restaurants": "Tourisme & Loisirs",
    "leisure": "Tourisme & Loisirs",
    "professional services": "Services", "commercial services": "Services",
    "trading companies": "Services", "distributors": "Services",
    "utilities": "Énergie", "energy": "Énergie", "oil and gas": "Énergie",
    "electric": "Énergie", "renewable": "Énergie", "power": "Énergie",
}

REGION_MAP = {
    "europe": "Europe", "northern america": "Amérique du Nord",
    "asia": "Asie", "latin america": "Amérique latine",
    "africa": "Afrique", "oceania": "Océanie", "mena": "Moyen-Orient",
}

FLAG_MAP = {
    "France": "🇫🇷", "Germany": "🇩🇪", "United Kingdom": "🇬🇧",
    "United States": "🇺🇸", "Japan": "🇯🇵", "China": "🇨🇳",
    "Denmark": "🇩🇰", "Sweden": "🇸🇪", "Netherlands": "🇳🇱",
    "Switzerland": "🇨🇭", "Belgium": "🇧🇪", "Italy": "🇮🇹",
    "Spain": "🇪🇸", "Norway": "🇳🇴", "Finland": "🇫🇮",
    "Canada": "🇨🇦", "Australia": "🇦🇺", "Brazil": "🇧🇷",
    "India": "🇮🇳", "South Korea": "🇰🇷", "Korea, Republic of": "🇰🇷",
    "Ireland": "🇮🇪", "Austria": "🇦🇹", "Portugal": "🇵🇹",
    "Luxembourg": "🇱🇺", "Mexico": "🇲🇽", "Singapore": "🇸🇬",
    "Poland": "🇵🇱", "Turkey": "🇹🇷", "Israel": "🇮🇱",
    "Colombia": "🇨🇴", "Chile": "🇨🇱", "Argentina": "🇦🇷",
    "Malaysia": "🇲🇾", "Indonesia": "🇮🇩",
}

SECTOR_EMOJI = {
    "Agroalimentaire": "🌾", "Distribution": "🛒", "Mode & Textile": "👗",
    "BTP & Immobilier": "🏗️", "Énergie": "⚡", "Transport": "🚛",
    "Tech & Numérique": "💻", "Finance": "🏦", "Industrie chimique": "🧪",
    "Industrie lourde": "⛏️", "Industrie manufacturière": "🏭",
    "Santé & Pharma": "💊", "Tourisme & Loisirs": "🏨", "Services": "📋",
}


def map_sector(raw):
    if not raw: return "Autre"
    low = raw.lower()
    for key, val in SECTOR_MAP.items():
        if key in low: return val
    return "Autre"

def map_region(raw):
    if not raw: return "Autre"
    low = str(raw).lower()
    for key, val in REGION_MAP.items():
        if key in low: return val
    return "Autre"

def extract_reduction(text):
    if not text: return None, None
    m12 = re.search(r'scope\s*1\s*(?:and|&)\s*(?:scope\s*)?2.*?(\d+(?:\.\d+)?)\s*%', text, re.I)
    mg = re.search(r'reduce\s+.*?(\d+(?:\.\d+)?)\s*%', text, re.I)
    m3 = re.search(r'scope\s*3.*?(\d+(?:\.\d+)?)\s*%', text, re.I)
    r12 = float(m12.group(1)) if m12 else (float(mg.group(1)) if mg else None)
    r3 = float(m3.group(1)) if m3 else None
    return r12, r3

def parse_year(val):
    if not val: return None
    try: return int(float(str(val).replace("FY", "").strip()))
    except: return None

def gen_id(name):
    return int(hashlib.md5(name.encode()).hexdigest()[:8], 16) % 100000 + 1000

def strip_html(text):
    if not text: return ""
    text = re.sub(r'<br\s*/?>', '\n', str(text))
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&lt;', '<', text)
    text = re.sub(r'&gt;', '>', text)
    text = re.sub(r'&#x27;', "'", text)
    text = re.sub(r'&quot;', '"', text)
    return text.strip()

def format_tonnes(val):
    try:
        t = float(val)
        if t >= 1000000: return f"{t/1000000:.1f} MtCO₂"
        if t >= 1000: return f"{t/1000:.0f} ktCO₂"
        return f"{t:.0f} tCO₂"
    except: return None


# ─── SBTi ───────────────────────────────

def build_sbti_entry(row):
    status = str(row.get("near_term_status") or "").strip()
    if status != "Targets set": return None
    name = str(row.get("company_name") or "").strip()
    if not name: return None

    country = str(row.get("location") or "").strip()
    region = map_region(row.get("region"))
    sector = map_sector(str(row.get("sector") or ""))
    org_type = str(row.get("organization_type") or "").strip()
    classification = str(row.get("near_term_target_classification") or "").strip()
    target_lang = str(row.get("full_target_language") or "").strip()
    nz_status = str(row.get("net_zero_status") or "").strip()
    target_year = parse_year(row.get("near_term_target_year"))
    nz_year = parse_year(row.get("net_zero_year"))
    r12, r3 = extract_reduction(target_lang)
    size = "PME" if org_type == "SME" else "Grande entreprise"
    flag = FLAG_MAP.get(country, "🏳️")
    logo = SECTOR_EMOJI.get(sector, "🏢")

    parts = []
    if r12: parts.append(f"-{r12:.0f}% d'émissions")
    if classification: parts.append(f"aligné {classification}")
    title = " — ".join(parts) if parts else "Objectifs SBTi validés"

    summary = ""
    if target_lang and len(target_lang) > 30:
        s = re.sub(r'https?://\S+\s*', '', target_lang)
        s = re.sub(r'This target was approved.*?(?:SMEs|enterprises)\.\s*', '', s, flags=re.DOTALL|re.I)
        s = re.sub(r'^Near-[Tt]erm [Tt]argets?:?\s*', '', s.strip())
        summary = (s[:300] + "...") if len(s) > 300 else s
    if not summary:
        summary = f"{name} a fait valider ses objectifs SBTi ({classification or 'Accord de Paris'}). Secteur : {sector}."

    stats = []
    if r12: stats.append({"value": f"-{r12:.0f}%", "label": "réduction Scope 1&2", "color": "#059669"})
    if r3: stats.append({"value": f"-{r3:.0f}%", "label": "réduction Scope 3", "color": "#D97706"})
    if classification: stats.append({"value": classification, "label": "alignement", "color": "#0284C7"})
    if nz_status == "Targets set" and nz_year:
        stats.append({"value": str(nz_year), "label": "objectif net zero", "color": "#7C3AED"})
    elif target_year:
        stats.append({"value": str(target_year), "label": "horizon cible", "color": "#7C3AED"})
    if not stats:
        stats.append({"value": "SBTi", "label": "validé", "color": "#0284C7"})

    scope = "Net Zero" if nz_status == "Targets set" else ("Scope 3" if r3 else "Scope 1 & 2")
    tt = f"SBTi validé — {classification or 'Accord de Paris'}"
    if target_year: tt += f" (horizon {target_year})"
    if nz_status == "Targets set" and nz_year: tt += f" — Net Zero {nz_year}"

    actions = []
    if r12: actions.append(f"Réduction de {r12:.0f}% des émissions Scope 1&2 d'ici {target_year or '?'}")
    if r3: actions.append(f"Réduction de {r3:.0f}% des émissions Scope 3")
    actions.append(f"Trajectoire alignée {classification or 'Accord de Paris'}")
    if nz_status == "Targets set": actions.append(f"Engagement Net Zero d'ici {nz_year or '?'}")

    return {
        "id": gen_id(name), "company": name, "title": title,
        "sector": sector, "region": region, "scope": scope,
        "source": "SBTi", "country": country or "?",
        "countryFlag": flag, "year": 2025, "logo": logo, "size": size,
        "summary": summary, "actions": actions, "stats": stats,
        "target": tt, "difficulty": "", "roi": "",
        "sourceUrl": "https://sciencebasedtargets.org/target-dashboard",
        "lastUpdated": datetime.now().strftime("%Y-%m-%d"),
        "contributors": ["Import automatique SBTi"],
        "verified": True, "autoImport": True,
    }


# ─── ADEME ───────────────────────────────

def load_ademe(filepath):
    """Load ADEME bilans GES from local Excel file."""
    print(f"  Fichier: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # 1. Load assessments
    ws = wb['assessments']
    headers = None
    assessments = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: headers = list(row); continue
        d = dict(zip(headers, row))
        if d.get('organization_type') == 'Entreprise':
            assessments[d['id']] = d

    # 2. Load action plans from texts
    ws2 = wb['texts']
    plans = defaultdict(dict)
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0: continue
        aid, key, val = row[0], row[1], row[2]
        if aid in assessments and val and len(str(val)) > 50:
            plans[aid][key] = strip_html(str(val))

    wb.close()
    print(f"  {len(assessments)} bilans entreprises, {len(plans)} avec plans d'action")
    return assessments, plans


def build_ademe_entries(assessments, plans):
    """Build project entries from ADEME data."""
    entries = []
    for aid, a in assessments.items():
        name = str(a.get('organization_name') or '').strip()
        if not name: continue

        try:
            s1 = float(a.get('total_scope_1') or 0)
            s2 = float(a.get('total_scope_2') or 0)
        except: continue
        if s1 + s2 < 100: continue  # Skip tiny emitters

        total = s1 + s2
        s3 = a.get('total_scope_3')
        staff = a.get('staff')
        year = a.get('reporting_year')
        red_s12 = a.get('reductions_scope_1_2')

        # Get action plans
        plan_texts = plans.get(aid, {})
        plan_s1 = plan_texts.get("Plan d'action Scope 1", "")
        plan_s2 = plan_texts.get("Plan d'action Scope 2", "")
        plan_s3 = plan_texts.get("Plan d'action Scope 3", "")
        description = plan_texts.get("Présentation de l'organisation", "")
        policy = plan_texts.get("Politique de développement durable", "")

        # Build actions list from plans
        actions = []
        for plan_name, plan_text in [("Scope 1", plan_s1), ("Scope 2", plan_s2), ("Scope 3", plan_s3)]:
            if plan_text and len(plan_text) > 30:
                # Extract first 2 sentences or 150 chars
                sentences = re.split(r'[.!?\n]', plan_text)
                meaningful = [s.strip() for s in sentences if len(s.strip()) > 20]
                if meaningful:
                    actions.append(f"{plan_name} : {meaningful[0][:150]}")

        if not actions:
            continue  # Skip entries without any action plan

        # Determine size
        try:
            staff_num = int(float(str(staff).replace(' ', ''))) if staff else 0
        except: staff_num = 0
        size = "PME" if staff_num < 500 else "Grande entreprise"

        # Summary
        summary_parts = []
        if description and len(description) > 50:
            summary_parts.append(description[:200])
        summary_parts.append(f"Émissions déclarées ({year}) : {format_tonnes(total)} (Scope 1&2).")
        if red_s12 and str(red_s12).strip() and str(red_s12) != 'None':
            try:
                red_val = float(red_s12)
                if red_val > 0:
                    summary_parts.append(f"Objectif de réduction : {format_tonnes(red_val)}.")
            except: pass
        summary = " ".join(summary_parts)
        if len(summary) > 350: summary = summary[:347] + "..."

        # Stats
        stats = [
            {"value": format_tonnes(total), "label": f"émissions Scope 1&2 ({year})", "color": "#DC2626"},
        ]
        if s3 and float(s3 or 0) > 0:
            stats.append({"value": format_tonnes(s3), "label": "Scope 3", "color": "#D97706"})
        if staff_num > 0:
            stats.append({"value": f"{staff_num:,}".replace(',', ' '), "label": "salariés", "color": "#0284C7"})
        if red_s12 and str(red_s12) != 'None':
            try:
                rv = float(red_s12)
                if rv > 0:
                    stats.append({"value": f"-{format_tonnes(rv)}", "label": "objectif réduction", "color": "#059669"})
            except: pass

        # Title
        title = f"Bilan GES déclaré — {format_tonnes(total)} (Scope 1&2)"

        entries.append({
            "id": gen_id(f"ADEME-{name}"),
            "company": name, "title": title,
            "sector": "Autre",  # ADEME doesn't categorize by sector
            "region": "Europe", "scope": "Scope 1 & 2",
            "source": "ADEME", "country": "France",
            "countryFlag": "🇫🇷", "year": int(year) if year else 2020,
            "logo": "📊", "size": size,
            "summary": summary, "actions": actions[:4], "stats": stats[:4],
            "target": f"Bilan GES réglementaire — année de reporting {year}",
            "difficulty": "", "roi": "",
            "sourceUrl": a.get('source_url') or "https://bilans-ges.ademe.fr",
            "lastUpdated": datetime.now().strftime("%Y-%m-%d"),
            "contributors": ["Import automatique ADEME"],
            "verified": True, "autoImport": True,
        })
    return entries


# ─── MAIN ───────────────────────────────

def main():
    print("=" * 50)
    print("Nouveaux Récits — Build Data")
    print("=" * 50)

    # 1. SBTi
    sbti_entries = []
    if os.path.exists(SBTI_LOCAL):
        print(f"\n1. SBTi ({SBTI_LOCAL})...")
        wb = openpyxl.load_workbook(SBTI_LOCAL, read_only=True, data_only=True)
        ws = wb.active
        headers = None
        total = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: headers = list(row); continue
            e = build_sbti_entry(dict(zip(headers, row)))
            if e: sbti_entries.append(e)
            total += 1
        wb.close()
        print(f"  {total} lignes → {len(sbti_entries)} avec objectifs validés")
    else:
        print(f"\n1. SBTi: fichier non trouvé ({SBTI_LOCAL})")

    # 2. ADEME
    ademe_entries = []
    if os.path.exists(ADEME_LOCAL):
        print(f"\n2. ADEME ({ADEME_LOCAL})...")
        assessments, plans = load_ademe(ADEME_LOCAL)
        ademe_entries = build_ademe_entries(assessments, plans)
        print(f"  {len(ademe_entries)} entreprises avec plans d'action")
    else:
        print(f"\n2. ADEME: fichier non trouvé ({ADEME_LOCAL})")

    # 3. Enrich SBTi French entries with ADEME data
    if ademe_entries:
        print(f"\n3. Enrichissement croisé SBTi ↔ ADEME...")
        ademe_by_name = {}
        for e in ademe_entries:
            key = e["company"].lower().strip()
            ademe_by_name[key] = e
        enriched = 0
        for e in sbti_entries:
            if e.get("country") == "France":
                key = e["company"].lower().strip()
                ad = ademe_by_name.get(key)
                if ad:
                    # Add ADEME emissions data to SBTi entry
                    for s in ad["stats"]:
                        if "émissions" in s["label"]:
                            e["stats"].append(s)
                            break
                    if ad.get("actions"):
                        e["actions"].extend(ad["actions"][:2])
                    e["source"] = "SBTi + ADEME"
                    enriched += 1
        print(f"  {enriched} fiches SBTi enrichies avec données ADEME")

    # 4. Editorial
    print(f"\n4. Fiches éditoriales...")
    editorial = []
    if os.path.exists(EDITORIAL_FILE):
        with open(EDITORIAL_FILE, "r", encoding="utf-8") as f:
            editorial = json.load(f)
        print(f"  {len(editorial)} fiches")

    # 5. Merge (editorial > SBTi > ADEME, no duplicates)
    print(f"\n5. Fusion...")
    names = {e["company"].lower().strip() for e in editorial}
    sbti_added = []
    for e in sbti_entries:
        k = e["company"].lower().strip()
        if k not in names:
            sbti_added.append(e)
            names.add(k)
    ademe_added = []
    for e in ademe_entries:
        k = e["company"].lower().strip()
        if k not in names:
            ademe_added.append(e)
            names.add(k)

    merged = editorial + sbti_added + ademe_added
    print(f"  {len(editorial)} éditoriales + {len(sbti_added)} SBTi + {len(ademe_added)} ADEME = {len(merged)} total")

    # 6. Write
    print(f"\n6. Écriture...")
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    # Stats
    pme = len([e for e in merged if e.get("size") == "PME"])
    nz = len([e for e in merged if e.get("scope") == "Net Zero"])
    fr = len([e for e in merged if e.get("country") == "France"])
    ademe_src = len([e for e in merged if "ADEME" in (e.get("source") or "")])
    print(f"\n✅ {len(merged)} projets")
    print(f"   {pme} PME | {nz} Net Zero | {fr} France | {ademe_src} ADEME")


if __name__ == "__main__":
    main()
